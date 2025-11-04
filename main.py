import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import openpyxl
import pandas as pd
import os
import pdfplumber
from pathlib import Path
import re

class PDFConverter:
    def __init__(self, pdf_file):
        self.pdf_file = pdf_file
        
    def extract_data_from_pdf(self):
        """Estrae i dati dall'ordine PDF con logica specifica per il formato GSD"""
        try:
            with pdfplumber.open(self.pdf_file) as pdf:
                all_data = []
                
                for page in pdf.pages:
                    # Prova prima con l'estrazione delle tabelle
                    tables = page.extract_tables()
                    
                    for table in tables:
                        for i, row in enumerate(table):
                            # Salta l'header della tabella
                            if i == 0 and ("Article Ref" in str(row) or "Cases Ordered" in str(row)):
                                continue
                            
                            # Cerca righe con il formato: numero, numero, numero
                            if (row and len(row) >= 3 and 
                                row[0] and row[1] and row[2] and
                                self._looks_like_article_data(row)):
                                clean_row = self._clean_row_data(row)
                                if clean_row:
                                    all_data.append(clean_row)
                    
                    # Se non ha trovato dati nelle tabelle, prova con l'estrazione del testo
                    if not all_data:
                        text = page.extract_text()
                        if text:
                            pdf_data = self._extract_from_text(text)
                            all_data.extend(pdf_data)
                
                return all_data
                
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile leggere il PDF: {str(e)}")
            return None
    
    def _looks_like_article_data(self, row):
        """Verifica se la riga sembra contenere dati di articoli"""
        if len(row) < 3:
            return False
        
        # Il primo campo dovrebbe essere un codice articolo (solo numeri)
        article_ref = str(row[0]).strip()
        if article_ref and re.match(r'^\d{5,}$', article_ref):  # Almeno 5 cifre
            return True
        return False
    
    def _clean_row_data(self, row):
        """Pulisce i dati della riga"""
        try:
            article_ref = str(row[0]).strip()
            cases_ordered = str(row[1]).strip().replace(',', '.')
            unit_qty = str(row[2]).strip().replace(',', '.')
            
            # Verifica che siano numeri validi
            if (re.match(r'^\d+$', article_ref) and
                re.match(r'^\d*\.?\d+$', cases_ordered) and
                re.match(r'^\d*\.?\d+$', unit_qty)):
                return [article_ref, float(cases_ordered), float(unit_qty)]
        except:
            pass
        return None
    
    def _extract_from_text(self, text):
        """Estrae dati dal testo del PDF"""
        data = []
        lines = text.split('\n')
        
        for line in lines:
            # Cerca pattern: numero (8+ cifre) seguito da numeri decimali
            match = re.search(r'(\d{8,})\s+(\d+\.?\d*)\s+(\d+\.?\d*)', line.strip())
            if match:
                article_ref = match.group(1)
                cases_ordered = match.group(2)
                unit_qty = match.group(3)
                data.append([article_ref, float(cases_ordered), float(unit_qty)])
        
        return data
    
    def pdf_to_excel(self):
        """Converte il PDF in un file Excel temporaneo"""
        data = self.extract_data_from_pdf()
        if not data:
            messagebox.showerror("Errore", "Nessun dato trovato nel PDF. Verifica il formato del file.")
            return None
        
        try:
            # Crea un nuovo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Order Data"
            
            # Intestazioni
            headers = ["Article Ref", "Cases Ordered", "Unit Qty"]
            ws.append(headers)
            
            # Aggiungi i dati
            for row in data:
                if len(row) >= 3:
                    ws.append([row[0], row[1], row[2]])
            
            # Formatta le colonne
            for col in range(1, 4):
                col_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15
            
            # Salva il file Excel temporaneo
            temp_file = os.path.join(os.path.dirname(self.pdf_file), 
                                   f"temp_conversion_{os.path.basename(self.pdf_file).replace('.pdf', '.xlsx')}")
            wb.save(temp_file)
            wb.close()
            
            messagebox.showinfo("PDF Convertito", f"PDF convertito con successo!\nTrovati {len(data)} articoli.\nEsempio: {data[0][0]} - {data[0][1]} - {data[0][2]}")
            return temp_file
            
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile convertire PDF in Excel: {str(e)}")
            return None

class SparConverter:
    def __init__(self, conversion_file, input_file):
        self.conversion_file = conversion_file
        self.input_file = input_file
        self.wb = None
        self.ws = None
        self.start_row = None
        
    def debug_data(self):
        """Mostra i dati per debug"""
        debug_info = f"File: {os.path.basename(self.input_file)}\n"
        debug_info += f"Righe totali: {self.ws.max_row}\n"
        debug_info += f"Colonne totali: {self.ws.max_column}\n"
        debug_info += f"Riga di partenza: {self.start_row}\n\n"
        
        debug_info += "PRIME 5 RIGHE:\n"
        for row in range(1, min(6, self.ws.max_row + 1)):
            row_data = []
            for col in range(1, min(6, self.ws.max_column + 1)):
                cell_value = self.ws.cell(row=row, column=col).value
                row_data.append(str(cell_value))
            debug_info += f"Riga {row}: {', '.join(row_data)}\n"
        
        return debug_info
    
    def load_workbook(self):
        """Carica il file Excel di input"""
        try:
            self.wb = openpyxl.load_workbook(self.input_file)
            self.ws = self.wb.active
            return True
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile caricare il file: {str(e)}")
            return False
    
    def pre_processing(self):
        """Esegue il pre-processing: rimuove merge, wrap text, etc."""
        # 1. Rimuovi tutti i merge
        merged_ranges = list(self.ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            self.ws.unmerge_cells(str(merged_range))
        
        # 2. Rimuovi wrap text da tutte le celle
        for row in self.ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=False)
        
        # 3. Imposta altezza uniforme di 15 per tutte le righe
        for row in range(1, self.ws.max_row + 1):
            self.ws.row_dimensions[row].height = 15
        
        # 4. Auto-adatta la larghezza di tutte le colonne
        for col in range(1, self.ws.max_column + 1):
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col)
            for row in range(1, self.ws.max_row + 1):
                try:
                    cell_value = self.ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[col_letter].width = adjusted_width
    
    def get_start_row(self):
        """Chiede all'utente la riga di partenza"""
        root = tk.Tk()
        root.withdraw()
        
        user_input = simpledialog.askstring(
            "Riga di Partenza", 
            "Inserisci il numero della riga di partenza (di solito 2 per file PDF convertiti):", 
            initialvalue="2"
        )
        
        root.destroy()
        
        if user_input is None or user_input == "":
            return None
        
        try:
            return int(user_input)
        except ValueError:
            messagebox.showerror("Errore", "Inserisci un numero valido!")
            return None
    
    def load_conversion_table(self):
        """Carica la tabella di conversione dal file SPAR CONVERSION.xlsm"""
        try:
            conversion_wb = openpyxl.load_workbook(self.conversion_file)
            conversion_ws = conversion_wb['Sheet1']
            
            # Crea un dizionario per la conversione (colonna B -> colonna C)
            conversion_dict = {}
            debug_info = "TABELLA DI CONVERSIONE (prime 10 righe):\n"
            
            for row in range(1, 131):  # Da riga 1 a 130
                key_cell = conversion_ws[f'B{row}']
                value_cell = conversion_ws[f'C{row}']
                if key_cell.value is not None and value_cell.value is not None:
                    conversion_dict[key_cell.value] = value_cell.value
                    if row <= 10:  # Mostra solo prime 10 righe per debug
                        debug_info += f"Riga {row}: {key_cell.value} -> {value_cell.value}\n"
            
            conversion_wb.close()
            
            # Mostra debug della tabella di conversione
            messagebox.showinfo("Debug Tabella Conversione", debug_info)
            return conversion_dict
            
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile caricare la tabella di conversione: {str(e)}")
            return None
    
    def apply_vlookup(self, conversion_dict):
        """Applica l'equivalente di VLOOKUP nella colonna C"""
        last_row = self.ws.max_row
        lookup_results = []
        
        for row in range(self.start_row, last_row + 1):
            try:
                lookup_value = self.ws[f'A{row}'].value
                original_value = lookup_value
                
                if lookup_value is not None:
                    # Converti a intero se possibile
                    try:
                        if isinstance(lookup_value, str):
                            lookup_value = lookup_value.strip()
                        lookup_int = int(lookup_value)
                        result = conversion_dict.get(lookup_int, 0)
                    except (ValueError, TypeError) as e:
                        result = 0
                else:
                    result = 0
                
                self.ws[f'C{row}'] = result
                lookup_results.append(f"Riga {row}: {original_value} -> {result}")
                
            except Exception as e:
                self.ws[f'C{row}'] = 0
                lookup_results.append(f"Riga {row}: ERRORE -> 0")
        
        # Mostra i risultati del lookup
        results_text = "\n".join(lookup_results[:10])  # Mostra prime 10 righe
        if len(lookup_results) > 10:
            results_text += f"\n... e altre {len(lookup_results) - 10} righe"
        
        messagebox.showinfo("Risultati VLOOKUP", results_text)
    
    def insert_column_and_apply_formula(self):
        """Inserisce una colonna tra C e D e applica la formula IF"""
        # Inserisce colonna D (dopo C)
        self.ws.insert_cols(4)
        
        last_row = self.ws.max_row
        
        # Codici speciali per le moltiplicazioni (come nel VBA originale)
        multiply_4_codes = [11005101, 11005102, 11005111, 11005112, 11005107, 11005113]
        multiply_3_codes = [11005382, 11005387]
        multiply_2_codes = [11004140, 11004141]
        
        calculation_results = []
        
        for row in range(self.start_row, last_row + 1):
            try:
                code = self.ws[f'C{row}'].value
                value_e = self.ws[f'E{row}'].value
                
                # Gestione valori None
                if value_e is None:
                    value_e = 0
                else:
                    try:
                        value_e = float(value_e)
                    except (ValueError, TypeError):
                        value_e = 0
                
                # Applica le moltiplicazioni come nel VBA originale
                if code in multiply_4_codes:
                    result = value_e * 4
                    multiplier = 4
                elif code in multiply_3_codes:
                    result = value_e * 3
                    multiplier = 3
                elif code in multiply_2_codes:
                    result = value_e * 2
                    multiplier = 2
                else:
                    result = value_e * 1
                    multiplier = 1
                
                self.ws[f'D{row}'] = result
                calculation_results.append(f"Riga {row}: Codice {code} x {multiplier} = {result}")
                
            except Exception as e:
                self.ws[f'D{row}'] = 0
                calculation_results.append(f"Riga {row}: ERRORE -> 0")
        
        # Mostra i risultati dei calcoli
        if calculation_results:
            calc_text = "\n".join(calculation_results[:10])
            if len(calculation_results) > 10:
                calc_text += f"\n... e altre {len(calculation_results) - 10} righe"
            messagebox.showinfo("Risultati Calcoli", calc_text)
    
    def delete_zero_rows(self):
        """Elimina le righe con 0 nella colonna C"""
        rows_to_delete = []
        
        # Trova le righe da eliminare
        for row in range(self.start_row, self.ws.max_row + 1):
            try:
                cell_value = self.ws[f'C{row}'].value
                if cell_value == 0 or cell_value == "0":
                    rows_to_delete.append(row)
            except:
                pass
        
        # Mostra quali righe verranno eliminate
        if rows_to_delete:
            delete_info = f"Righe da eliminare (con 0 in colonna C): {rows_to_delete}"
            messagebox.showinfo("Debug Eliminazione", delete_info)
        
        # Elimina le righe dalla fine per evitare problemi con gli indici
        deleted_count = 0
        for row in sorted(rows_to_delete, reverse=True):
            self.ws.delete_rows(row)
            deleted_count += 1
        
        return deleted_count
    
    def convert(self, is_pdf_conversion=False):
        """Esegue l'intero processo di conversione"""
        if not self.load_workbook():
            return False
        
        # DEBUG: Mostra i dati prima della conversione
        debug_info = self.debug_data()
        messagebox.showinfo("Debug Dati Input", debug_info)
        
        # PRE-STEP: Formattazione iniziale
        self.pre_processing()
        
        # INPUT: Chiedi all'utente la riga di partenza
        self.start_row = self.get_start_row()
        if self.start_row is None:
            return False
        
        # Verifica che la riga di partenza sia valida
        if self.start_row > self.ws.max_row:
            messagebox.showerror("Errore", "La riga di partenza è oltre l'ultima riga con dati!")
            return False
        
        # Carica la tabella di conversione
        conversion_dict = self.load_conversion_table()
        if conversion_dict is None:
            return False
        
        # PRIMO STEP: Applica VLOOKUP nella colonna C
        self.apply_vlookup(conversion_dict)
        
        # SECONDO STEP: Inserisce una colonna tra C e D
        self.insert_column_and_apply_formula()
        
        # TERZO STEP: Elimina righe con 0 nella colonna C
        deleted_rows = self.delete_zero_rows()
        
        # QUARTO STEP: Ri-applica auto-fit alle colonne
        for col in range(1, self.ws.max_column + 1):
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col)
            for row in range(1, self.ws.max_row + 1):
                try:
                    cell_value = self.ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[col_letter].width = adjusted_width
        
        # Salva il file convertito
        if is_pdf_conversion:
            # Usa il nome originale del PDF
            original_pdf_name = os.path.basename(self.input_file).replace('temp_conversion_', '').replace('.xlsx', '')
            output_file = os.path.join(os.path.dirname(self.input_file), f"{original_pdf_name}_CONVERTITO.xlsx")
        else:
            base_name = os.path.splitext(self.input_file)[0]
            output_file = f"{base_name}_CONVERTITO.xlsx"
        
        try:
            self.wb.save(output_file)
            self.wb.close()
            
            # DEBUG: Controlla se il file finale ha dati
            if os.path.exists(output_file):
                final_wb = openpyxl.load_workbook(output_file)
                final_ws = final_wb.active
                final_row_count = final_ws.max_row
                final_wb.close()
                
                final_info = f"File salvato: {output_file}\nRighe nel file finale: {final_row_count}"
                messagebox.showinfo("Debug File Finale", final_info)
            
            # Messaggio di completamento
            messagebox.showinfo(
                "Automazione Completata!",
                f"Conversione terminata con successo!\n\n"
                f"Riga di partenza: {self.start_row}\n"
                f"Righe eliminate: {deleted_rows}\n"
                f"File salvato come: {os.path.basename(output_file)}\n"
                f"Percorso: {output_file}"
            )
            
            # Apri la cartella contenente il file
            os.startfile(os.path.dirname(output_file))
            return True
            
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile salvare il file: {str(e)}")
            return False

def select_file(title, file_types):
    """Seleziona un file tramite dialog"""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=file_types)
    root.destroy()
    return file_path

def main():
    try:
        # Seleziona il file di conversione SPAR
        conversion_file = select_file(
            "Seleziona il file SPAR CONVERSION.xlsm",
            [("Excel files", "*.xlsm"), ("All files", "*.*")]
        )
        
        if not conversion_file:
            return
        
        # Chiedi all'utente se vuole convertire PDF o usare Excel
        root = tk.Tk()
        root.withdraw()
        
        choice = messagebox.askquestion(
            "Tipo di File",
            "Vuoi convertire un file PDF o un file Excel?\n\n"
            "• Sì = Converti PDF\n"
            "• No = Usa file Excel esistente",
            icon='question'
        )
        root.destroy()
        
        input_file = None
        is_pdf_conversion = False
        
        if choice == 'yes':
            # Conversione PDF
            pdf_file = select_file(
                "Seleziona il file PDF da convertire",
                [("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            
            if pdf_file:
                messagebox.showinfo("Conversione PDF", "Sto convertendo il PDF in Excel...")
                pdf_converter = PDFConverter(pdf_file)
                input_file = pdf_converter.pdf_to_excel()
                is_pdf_conversion = True
                
                if not input_file:
                    return
        else:
            # File Excel esistente
            input_file = select_file(
                "Seleziona il file Excel da convertire",
                [("Excel files", "*.xlsx"), ("Excel files", "*.xls"), ("All files", "*.*")]
            )
        
        if not input_file:
            return
        
        # Esegue la conversione SPAR
        converter = SparConverter(conversion_file, input_file)
        success = converter.convert(is_pdf_conversion)
        
        # Pulisci file temporaneo se era una conversione PDF
        if is_pdf_conversion and input_file and os.path.exists(input_file):
            try:
                os.remove(input_file)
                print(f"File temporaneo cancellato: {input_file}")
            except Exception as e:
                print(f"Non è stato possibile cancellare il file temporaneo: {e}")
        
        if not success:
            messagebox.showerror("Errore", "La conversione non è stata completata.")
            
    except Exception as e:
        messagebox.showerror("Errore Critico", f"Si è verificato un errore: {str(e)}")

if __name__ == "__main__":
    main()
