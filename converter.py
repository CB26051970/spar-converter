import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import simpledialog, messagebox
import os

class SparConverter:
    def __init__(self, conversion_file, input_file):
        self.conversion_file = conversion_file
        self.input_file = input_file
        self.ws = None
        self.start_row = None
        
    def load_workbook(self):
        """Carica il file Excel di input"""
        try:
            self.wb = openpyxl.load_workbook(self.input_file)
            self.ws = self.wb.active
            return True
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile caricare il file: {str(e)}")
            return False
    
    def pre_processing(self):
        """Esegue il pre-processing: rimuove merge, wrap text, etc."""
        # Rimuovi tutti i merge
        for merged_range in list(self.ws.merged_cells.ranges):
            self.ws.unmerge_cells(str(merged_range))
        
        # Rimuovi wrap text
        for row in self.ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=False)
        
        # Imposta altezza uniforme delle righe
        for row in range(1, self.ws.max_row + 1):
            self.ws.row_dimensions[row].height = 15
        
        # Auto-adatta le colonne
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_start_row(self):
        """Chiede all'utente la riga di partenza"""
        root = tk.Tk()
        root.withdraw()  # Nasconde la finestra principale
        
        user_input = simpledialog.askstring(
            "Riga di Partenza", 
            "Inserisci il numero della riga di partenza (es. 5 o 6):", 
            initialvalue="6"
        )
        
        root.destroy()
        
        if user_input is None:
            return None
        
        try:
            return int(user_input)
        except ValueError:
            messagebox.showerror("Errore", "Inserisci un numero valido!")
            return None
    
    def apply_vlookup(self, conversion_df):
        """Applica l'equivalente di VLOOKUP nella colonna C"""
        last_row = self.ws.max_row
        
        for row in range(self.start_row, last_row + 1):
            try:
                lookup_value = self.ws[f'A{row}'].value
                if lookup_value is not None:
                    # Cerca il valore nella tabella di conversione
                    match = conversion_df[conversion_df.iloc[:, 0] == int(lookup_value)]
                    if not match.empty:
                        self.ws[f'C{row}'] = match.iloc[0, 1]
                    else:
                        self.ws[f'C{row}'] = 0
                else:
                    self.ws[f'C{row}'] = 0
            except (ValueError, TypeError):
                self.ws[f'C{row}'] = 0
    
    def insert_column_and_apply_formula(self):
        """Inserisce una colonna tra C e D e applica la formula IF"""
        # Inserisce colonna D
        self.ws.insert_cols(4)  # Inserisce dopo la colonna C (indice 4 = colonna D)
        
        last_row = self.ws.max_row
        
        # Codici speciali per le moltiplicazioni
        multiply_4_codes = [11005101, 11005102, 11005111, 11005112, 11005107, 11005113]
        multiply_3_codes = [11005382, 11005387]
        multiply_2_codes = [11004140, 11004141]
        
        for row in range(self.start_row, last_row + 1):
            try:
                code = self.ws[f'C{row}'].value
                value_e = self.ws[f'E{row}'].value
                
                if value_e is None:
                    value_e = 0
                
                if code in multiply_4_codes:
                    result = value_e * 4
                elif code in multiply_3_codes:
                    result = value_e * 3
                elif code in multiply_2_codes:
                    result = value_e * 2
                else:
                    result = value_e * 1
                
                self.ws[f'D{row}'] = result
            except (ValueError, TypeError):
                self.ws[f'D{row}'] = 0
    
    def delete_zero_rows(self):
        """Elimina le righe con 0 nella colonna C"""
        rows_to_delete = []
        
        for row in range(self.start_row, self.ws.max_row + 1):
            if self.ws[f'C{row}'].value == 0:
                rows_to_delete.append(row)
        
        # Elimina le righe dalla fine per evitare problemi con gli indici
        for row in sorted(rows_to_delete, reverse=True):
            self.ws.delete_rows(row)
        
        return len(rows_to_delete)
    
    def load_conversion_table(self):
        """Carica la tabella di conversione dal file SPAR CONVERSION.xlsm"""
        try:
            conversion_wb = openpyxl.load_workbook(self.conversion_file)
            conversion_ws = conversion_wb['Sheet1']
            
            # Legge i dati dalla colonna B1 a C130
            conversion_data = []
            for row in range(1, 131):
                cell_b = conversion_ws[f'B{row}'].value
                cell_c = conversion_ws[f'C{row}'].value
                if cell_b is not None and cell_c is not None:
                    conversion_data.append([cell_b, cell_c])
            
            conversion_df = pd.DataFrame(conversion_data)
            conversion_wb.close()
            return conversion_df
            
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile caricare la tabella di conversione: {str(e)}")
            return None
    
    def convert(self, output_file=None):
        """Esegue l'intero processo di conversione"""
        if not self.load_workbook():
            return False
        
        # Pre-processing
        self.pre_processing()
        
        # Ottiene la riga di partenza
        self.start_row = self.get_start_row()
        if self.start_row is None:
            return False
        
        # Carica la tabella di conversione
        conversion_df = self.load_conversion_table()
        if conversion_df is None:
            return False
        
        # Applica VLOOKUP
        self.apply_vlookup(conversion_df)
        
        # Inserisce colonna e applica formula
        self.insert_column_and_apply_formula()
        
        # Elimina righe con zero
        deleted_rows = self.delete_zero_rows()
        
        # Ri-applica auto-fit
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salva il file
        if output_file is None:
            base_name = os.path.splitext(self.input_file)[0]
            output_file = f"{base_name}_converted.xlsx"
        
        self.wb.save(output_file)
        self.wb.close()
        
        # Mostra messaggio di completamento
        messagebox.showinfo(
            "Completato",
            f"Automazione completata!\n"
            f"Riga di partenza: {self.start_row}\n"
            f"Righe eliminate: {deleted_rows}\n"
            f"File salvato come: {output_file}"
        )
        
        return True
